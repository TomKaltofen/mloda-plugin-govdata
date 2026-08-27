"""Destatis GV-ISys change-file loader (Gebietsstand history, ADR 0006).

Each year's "Namens-Grenz-Aenderung" workbook uses a two-level merged-cell header
whose row offset is not something a single sample year can pin reliably across the
full 1990-2024 span. Data rows are found by their own shape instead: the first
column holds a change id (``<land>/<year>/<seq>-<letter>``, e.g. ``03/2016/0006-R``),
a pattern no header or blank row matches.
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from datetime import date
from typing import Any

import openpyxl

from mloda_plugin_govdata.feature_groups.govdata.core.cache import DownloadCache

from .download import fetch_pinned
from .sources import gv_isys_source

_CHANGE_ID = re.compile(r"^\d{2}/\d{4}/\d+-[A-Z]$")


@dataclass(frozen=True)
class GvIsysChange:
    change_id: str
    level: str  # "Kreis" or "Gemeinde"
    from_rs: str
    from_ags: str
    from_name: str
    change_type: str
    to_rs: str
    to_ags: str
    to_name: str
    effective_date_legal: date  # "juristisch"
    effective_date_statistical: date  # "statistisch"


def _parse_date(value: Any) -> date:
    """Parses a German ``dd.mm.yyyy`` calendar date. No time-of-day, so no timezone applies."""
    if isinstance(value, date):
        return value
    day, month, year = str(value).strip().split(".")
    return date(int(year), int(month), int(day))


def parse_gv_isys_workbook(path: str | os.PathLike[str]) -> list[GvIsysChange]:
    """Parses one year's GV-ISys "Namens-Grenz-Aenderung" workbook into change rows.

    A merger or split reads as several rows sharing one ``change_id`` (one row per
    donating/dissolving unit).
    """
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    changes: list[GvIsysChange] = []
    for cells in sheet.iter_rows(values_only=True):
        change_id = cells[0]
        if not isinstance(change_id, str) or not _CHANGE_ID.match(change_id):
            continue  # title, blank, or header row; data rows are self-describing
        (
            _change_id,
            level,
            from_rs,
            from_ags,
            from_name,
            change_type,
            _area_sqm,
            _population,
            to_rs,
            to_ags,
            to_name,
            legal_date,
            statistical_date,
        ) = cells
        changes.append(
            GvIsysChange(
                change_id=change_id,
                level=str(level).strip(),
                from_rs=str(from_rs).strip(),
                from_ags=str(from_ags).strip(),
                from_name=str(from_name).strip(),
                change_type=str(change_type).strip(),
                to_rs=str(to_rs).strip(),
                to_ags=str(to_ags).strip(),
                to_name=str(to_name).strip(),
                effective_date_legal=_parse_date(legal_date),
                effective_date_statistical=_parse_date(statistical_date),
            )
        )
    return changes


def load_gv_isys_changes(
    year: int, cache: DownloadCache, *, sha256: str | None = None, revalidate: bool = False
) -> list[GvIsysChange]:
    """Fetches (offline-cache-first) and parses one year's GV-ISys change rows.

    ``sha256`` pins the download when known (2016 is pinned via ``sources.GV_ISYS_2016_SHA256``);
    other years fetch unpinned.
    """
    source = gv_isys_source(year, sha256=sha256)
    path = fetch_pinned(cache, source, revalidate=revalidate)
    return parse_gv_isys_workbook(path)
