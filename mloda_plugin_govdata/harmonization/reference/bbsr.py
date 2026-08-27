"""BBSR Umsteigeschluessel Kreise loader (proportional re-basing keys, ADR 0006)."""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from typing import Any

import openpyxl

from mloda_plugin_govdata.feature_groups.govdata.core.cache import DownloadCache
from mloda_plugin_govdata.harmonization.keys import repair_bbsr_kreis_key

from .download import fetch_pinned
from .sources import BBSR_KREISE

_SHEET_NAME = re.compile(r"^(\d{4})-(\d{4})$")


@dataclass(frozen=True)
class UmsteigeschluesselRow:
    from_year: int
    to_year: int
    source_key: str
    source_name: str
    area_share: float
    population_share: float
    employee_share: float
    area_km2: float
    population_thousands: float
    svb_thousands: float  # sozialversicherungspflichtig Beschaeftigte (employees liable for social insurance)
    target_key: str
    target_name: str


def _parse_sheet(sheet: Any, from_year: int, to_year: int) -> list[UmsteigeschluesselRow]:
    rows: list[UmsteigeschluesselRow] = []
    row_iter = sheet.iter_rows(values_only=True)
    next(row_iter)  # header
    for cells in row_iter:
        if cells[0] is None:  # trailing all-empty row ends the sheet's data
            break
        src, src_name, area_share, pop_share, emp_share, area, pop, svb, tgt, tgt_name = cells
        rows.append(
            UmsteigeschluesselRow(
                from_year=from_year,
                to_year=to_year,
                source_key=repair_bbsr_kreis_key(src),
                source_name=str(src_name),
                area_share=float(area_share),
                population_share=float(pop_share),
                employee_share=float(emp_share),
                area_km2=float(area),
                population_thousands=float(pop),
                svb_thousands=float(svb),
                target_key=repair_bbsr_kreis_key(tgt),
                target_name=str(tgt_name),
            )
        )
    return rows


def parse_bbsr_kreise_workbook(path: str | os.PathLike[str]) -> list[UmsteigeschluesselRow]:
    """Parses every year-pair sheet of a BBSR Kreise Umsteigeschluessel workbook.

    The full file has 34 sheets, 1990 to 2024, one per consecutive year pair named
    ``<y>-<y+1>``; direction is old to new (forward), read from the sheet name. Does
    not validate per-key share sums: at least one sheet carries a known upstream defect
    where split shares land on identity rows instead of a transfer row (see the fixture
    ``NOTICE``); asserting and raising on that is the re-basing loader's job, not this one's.
    """
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows: list[UmsteigeschluesselRow] = []
    for sheet_name in workbook.sheetnames:
        match = _SHEET_NAME.match(sheet_name)
        if match is None:
            continue  # not a year-pair sheet; skip defensively rather than raise
        from_year, to_year = int(match.group(1)), int(match.group(2))
        rows.extend(_parse_sheet(workbook[sheet_name], from_year, to_year))
    return rows


def load_bbsr_kreise(cache: DownloadCache, *, revalidate: bool = False) -> list[UmsteigeschluesselRow]:
    """Fetches (offline-cache-first) and parses the BBSR Kreise Umsteigeschluessel."""
    path = fetch_pinned(cache, BBSR_KREISE, revalidate=revalidate)
    return parse_bbsr_kreise_workbook(path)
