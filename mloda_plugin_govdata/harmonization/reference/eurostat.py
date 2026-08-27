"""Eurostat NUTS/LAU reference-table loaders (ADR 0006)."""

from __future__ import annotations

import os
import re
from dataclasses import dataclass

import openpyxl

from mloda_plugin_govdata.feature_groups.govdata.core.cache import DownloadCache

from .download import fetch_pinned
from .sources import EUROSTAT_LAU_NUTS, EUROSTAT_NUTS_CORRESPONDENCE

_LAST_UPDATE = re.compile(r"last update (\d{2}/\d{2}/\d{4}).*based on (NUTS \d+ and LAU \d+)")


@dataclass(frozen=True)
class NutsCorrespondenceOverview:
    """The small DE summary row of Eurostat's "Correspondence table" (edition overview only).

    Not the crosswalk used for mapping: see ADR 0006, Edition identity. This table's own
    ``edition_label`` names the NUTS/LAU edition it was drawn from, which lags the one
    :func:`load_lau_nuts_de` uses (``nuts_version="2024"``); mismatch is expected, not a bug.
    """

    edition_label: str
    last_update: str
    laender: int
    regierungsbezirke: int
    kreise: int
    gemeinden: int


def parse_nuts_correspondence_workbook(path: str | os.PathLike[str]) -> NutsCorrespondenceOverview:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    last_update = ""
    edition_label = ""
    counts: tuple[int, int, int, int] | None = None
    for cells in sheet.iter_rows(values_only=True):
        first = cells[0]
        if isinstance(first, str):
            match = _LAST_UPDATE.search(first)
            if match is not None:
                last_update, edition_label = match.group(1), match.group(2)
        if first == "DE":
            counts = (int(cells[2]), int(cells[4]), int(cells[6]), int(cells[8]))
    if counts is None:
        raise ValueError("no 'DE' row found in the Eurostat NUTS correspondence table")
    return NutsCorrespondenceOverview(
        edition_label=edition_label,
        last_update=last_update,
        laender=counts[0],
        regierungsbezirke=counts[1],
        kreise=counts[2],
        gemeinden=counts[3],
    )


def load_nuts_correspondence_overview(cache: DownloadCache, *, revalidate: bool = False) -> NutsCorrespondenceOverview:
    """Fetches (offline-cache-first) and parses the Eurostat NUTS correspondence overview."""
    path = fetch_pinned(cache, EUROSTAT_NUTS_CORRESPONDENCE, revalidate=revalidate)
    return parse_nuts_correspondence_workbook(path)


@dataclass(frozen=True)
class LauNutsRow:
    period: int
    nuts3: str
    lau_code: str
    lau_name: str
    change: str
    population: int | None
    total_area_m2: int | None
    degurba: int | None
    coastal_area: bool


def parse_lau_nuts_de_workbook(path: str | os.PathLike[str]) -> list[LauNutsRow]:
    """Parses the Eurostat LAU-to-NUTS correspondence workbook, Germany sheet only.

    NUTS 2024 / LAU 2025 edition (ADR 0006, not the newer-labelled "2027" summary
    table from :func:`load_nuts_correspondence_overview`). Germany's row in the
    source file's own Overview sheet (not loaded here) is marked fully validated
    across every column; among all EU-27 countries only Cyprus carries a
    partial-validation caveat there (2018 FUA commuting data), so this DE-only
    loader never needs to surface a per-row validation flag.
    """
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["DE"]

    row_iter = sheet.iter_rows(values_only=True)
    next(row_iter)  # header: PERIOD, NUTS3, LAU CODE, EU LAU CODE, LAU NAME NATIONAL, ...
    rows: list[LauNutsRow] = []
    for cells in row_iter:
        if cells[0] is None:  # trailing all-empty row ends the sheet's data
            break
        period, nuts3, lau_code_cell, _eu_lau_code, lau_name, _lau_name_latin, change = cells[:7]
        population, total_area_m2, degurba, coastal_area = cells[7:11]
        lau_code = lau_code_cell if isinstance(lau_code_cell, str) else str(int(lau_code_cell)).zfill(8)
        rows.append(
            LauNutsRow(
                period=int(period),
                nuts3=str(nuts3),
                lau_code=lau_code,
                lau_name=str(lau_name),
                change=str(change),
                population=None if population is None else int(population),
                total_area_m2=None if total_area_m2 is None else int(total_area_m2),
                degurba=None if degurba is None else int(degurba),
                coastal_area=bool(coastal_area),
            )
        )
    return rows


def load_lau_nuts_de(cache: DownloadCache, *, revalidate: bool = False) -> list[LauNutsRow]:
    """Fetches (offline-cache-first) and parses the Eurostat LAU-to-NUTS Germany sheet."""
    path = fetch_pinned(cache, EUROSTAT_LAU_NUTS, revalidate=revalidate)
    return parse_lau_nuts_de_workbook(path)
